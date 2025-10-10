import re
import random
import warnings
import datetime
import calendar
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
import xlwings as xw
import pandas as pd
import pyodbc

# ---------- SUPPRESS WARNINGS ----------
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------- CONFIG ----------
TARGET_SHEETS = ["AKUN", "WAMA", "GALAXEA"]
TARGET_RED = "FFC00000"
HEADERS = ["Event", "Resource", "Configuration", "Date", "Start Time", "End Time", "Capacity", "Reference"]
HEADER_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FONT = Font(bold=True)
ROW_BOLD_FONT = Font(bold=True)
YEAR_FOR_OUTPUT = datetime.datetime.now().year
OFF_CODES = {"AO", "OD", "AL", "SK", "PH", "AB", "TL", "DO", "OF", "CV", "CVO", "OS"}
MONTH_MAP = {
    "january": 1, "jan": 1, "february": 2, "feb": 2, "march": 3, "mar": 3,
    "april": 4, "apr": 4, "may": 5, "june": 6, "jun": 6, "july": 7, "jul": 7,
    "august": 8, "aug": 8, "september": 9, "sep": 9, "sept": 9,
    "october": 10, "oct": 10, "november": 11, "nov": 11, "december": 12, "dec": 12
}
MONTH_SHORT = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun", 7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}
RESORT_SHORT_KEYWORDS = {
    "ST": ["st.regis", "st regis", "stregis", "st"],
    "NJ": ["nujuma", "nujuma resort", "nj"],
    "TB": ["turtle bay", "turtlebay", "tb"],
    "DR": ["desert rock", "desertrock", "dr"],
    "SS": ["six senses", "sixsenses", "ss"],
    "SH": ["shebara", "sh", "shebarah"],
    "ED": ["edition", "ed"],
    "MV": ["maravel", "mv"],
    "SLS": ["sls resort", "sls", "slsresort"],
    "IC": ["intercontinental", "ic"],
    "AM": ["amaala", "am"]
}

# ---------- HELPER FUNCTIONS ----------
def safe_str(v):
    return "" if v is None else str(v).strip()

def get_rgb(cell):
    if cell is None:
        return None
    fill = getattr(cell, "fill", None)
    if not fill:
        return None
    color = getattr(fill, "start_color", None)
    if color is None:
        return None
    if hasattr(color, "rgb") and color.rgb:
        return str(color.rgb).upper()
    return None

def is_red(cell):
    rgb = get_rgb(cell)
    return rgb == TARGET_RED or (rgb and rgb.endswith(TARGET_RED[-6:]))

def parse_month_to_num(month_value):
    if month_value is None:
        return None
    s = str(month_value).strip()
    if not s:
        return None
    if s.isdigit() and 1 <= int(s) <= 12:
        return int(s)
    key = s.lower()
    if key in MONTH_MAP:
        return MONTH_MAP[key]
    for name, num in MONTH_MAP.items():
        if name in key:
            return num
    m = re.search(r"\b(1[0-2]|0?[1-9])\b", s)
    if m:
        return int(m.group(1))
    return None

def add_headers(ws):
    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

def clean_instructor_name(name):
    if not name:
        return None
    return re.sub(r"\s*\(.*?\)\s*", "", str(name)).strip()

def get_light_fill():
    colors = [
        "FFFFE5CC", "FFE5FFCC", "FFCCFFE5", "FFCCE5FF",
        "FFFFCCFF", "FFE5CCFF", "FFFFCCCC", "FFCCFFFF"
    ]
    hexcolor = random.choice(colors)
    return PatternFill(start_color=hexcolor, end_color=hexcolor, fill_type="solid")

def get_resort_short_from_name(resort_name):
    if not resort_name:
        return None
    s = resort_name.lower()
    for short in RESORT_SHORT_KEYWORDS.keys():
        if s == short.lower():
            return short
    for short, keywords in RESORT_SHORT_KEYWORDS.items():
        for kw in keywords:
            if kw in s:
                return short
    tokens = re.findall(r"[A-Za-z0-9]+", s)
    for t in tokens:
        t_up = t.upper()
        if t_up in RESORT_SHORT_KEYWORDS:
            return t_up
    return None

def days_in_month(year, month):
    return calendar.monthrange(year, month)[1]

def get_last_data_row(ws, col_idx):
    last = ws.max_row
    while last > 1:
        val = ws.cell(row=last, column=col_idx).value
        if val is not None and str(val).strip() != "":
            return last
        last -= 1
    return 1

# ---------- STAFF & ROSTER PRELOAD ----------
def preload_staff(staff_file, status_callback=None, progress_state=None):
    wb = load_workbook(staff_file, data_only=True)
    result = {}
    # map workbook sheet names case-insensitively using UPPER
    staff_sheet_map = {s.upper(): s for s in wb.sheetnames}
    for target in TARGET_SHEETS:
        actual_sheet = staff_sheet_map.get(target.upper())
        if not actual_sheet:
            continue
        ws = wb[actual_sheet]
        headers = [safe_str(c.value).lower() for c in ws[1]]
        try:
            pri_idx = headers.index("priority") + 1
        except ValueError:
            pri_idx = 1
        instr_start_col = pri_idx + 1
        sheet_map = {}
        for col in range(instr_start_col, ws.max_column + 1):
            instr_name = clean_instructor_name(safe_str(ws.cell(row=1, column=col).value))
            if not instr_name:
                continue
            last_row_for_col = get_last_data_row(ws, col)
            r = 2
            while r <= last_row_for_col:
                if progress_state is not None and status_callback is not None:
                    progress_state['current'] += 1
                    status_callback(progress_state['current'], progress_state['total'], f"Preloading Staff {target} row {r}")
                val_cell = ws.cell(row=r, column=col)
                val = safe_str(val_cell.value)
                if not val:
                    break
                if is_red(val_cell):
                    r += 1
                    continue
                # normalize activity/config key to lower-case (staff specialty text)
                key = val.strip().lower()
                sheet_map.setdefault(key, []).append(instr_name)
                r += 1
        result[target.upper()] = sheet_map
    return result

def preload_roster(roster_file, status_callback=None, progress_state=None):
    """
    Returns availability, off_days
    status_callback(current, total, message) is optional.
    progress_state: dict with keys 'current' and 'total' to update row-level progress.
    """
    wb = load_workbook(roster_file, data_only=True)
    availability = {}
    off_days = {}
    for sheet_name in wb.sheetnames:
        # update progress for the header row scanning as well (approx)
        if progress_state is not None and status_callback is not None:
            progress_state['current'] += 1
            status_callback(progress_state['current'], progress_state['total'], f"Preloading Roster sheet {sheet_name}")
        month_num = parse_month_to_num(sheet_name)
        if month_num is None:
            month_num = parse_month_to_num(sheet_name.split()[0])
        if month_num is None:
            continue
        ws = wb[sheet_name]
        team_cell = None
        for r in range(1, ws.max_row + 1):
            # update progress when scanning to find team members header
            if progress_state is not None and status_callback is not None:
                progress_state['current'] += 1
                status_callback(progress_state['current'], progress_state['total'], f"Preloading Roster {sheet_name} scanning row {r}")
            for c in range(1, ws.max_column + 1):
                val = safe_str(ws.cell(row=r, column=c).value).strip().lower()
                if val == "team members name":
                    team_cell = (r, c)
                    break
            if team_cell:
                break
        if not team_cell:
            continue
        team_row, team_col = team_cell
        date_cols = []
        for c in range(team_col + 1, ws.max_column + 1):
            hdr = ws.cell(row=team_row, column=c).value
            if hdr is None:
                break
            try:
                dnum = int(hdr)
                if 1 <= dnum <= 31:
                    date_cols.append((c, dnum))
            except:
                try:
                    dnum = int(float(hdr))
                    if 1 <= dnum <= 31:
                        date_cols.append((c, dnum))
                except:
                    break
        # use true last row for team column
        last_row_for_team = get_last_data_row(ws, team_col)
        r = team_row + 1
        while r <= last_row_for_team:
            # per-row progress update
            if progress_state is not None and status_callback is not None:
                progress_state['current'] += 1
                status_callback(progress_state['current'], progress_state['total'], f"Preloading Roster {sheet_name} row {r}")
            name_cell = ws.cell(row=r, column=team_col)
            name = clean_instructor_name(safe_str(name_cell.value))
            if not name:
                break
            for (c, day_num) in date_cols:
                val = safe_str(ws.cell(row=r, column=c).value)
                if not val:
                    continue
                parts = re.split(r"[,/;|\n]+", val)
                for p in parts:
                    p_clean = p.strip()
                    if not p_clean:
                        continue
                    if p_clean.upper() in OFF_CODES:
                        off_days.setdefault((month_num, int(day_num)), {})[name] = p_clean.upper()
                        continue
                    short = None
                    if p_clean.upper() in RESORT_SHORT_KEYWORDS:
                        short = p_clean.upper()
                    else:
                        short = get_resort_short_from_name(p_clean)
                    if not short:
                        for s_code, keys in RESORT_SHORT_KEYWORDS.items():
                            for kw in keys:
                                if kw in p_clean.lower():
                                    short = s_code
                                    break
                            if short:
                                break
                    if not short:
                        short = p_clean.upper()
                    key = (month_num, int(day_num), short)
                    availability.setdefault(key, set()).add(name)
            r += 1
    return availability, off_days

def build_event_names(activity, resort, all_resorts_for_activity):
    out = []
    if not activity:
        return out
    # Display-only: always show raw activity name
    out.append(activity)
    return out

# ---------- EXCEL BOOKABLE HOURS ----------
def get_bookable_hours(ws, row, col):
    rng = ws.range((row, col))
    values = []
    try:
        dv = rng.api.Validation
        if dv and dv.Formula1:
            ref = dv.Formula1.lstrip("=")
            sheet, addr = ref.split("!")
            sheet = sheet.strip("'")
            avail_ws = ws.book.sheets[sheet]
            for v in avail_ws.range(addr):
                if v.value:
                    values.append(str(v.value).strip())
    except Exception as e:
        print(f"⚠️ Could not fetch dropdown for {getattr(ws,'name','unknown')}!{row},{col}: {e}")
    return values

# ---------- GENERATE OUTPUT ----------
def generate_output(events_file, staff_file, roster_file, output_file, status_callback=None):
    """
    generate_output(..., status_callback=None)
    status_callback(current, total, message) - optional; called during row processing.
    """
    # We will compute a conservative total rows to process across the three workbooks
    # so that the GUI progress bar can be determinate.
    total_rows = 0
    try:
        wb_src_tmp = load_workbook(events_file, data_only=True)
        for sheet_name in wb_src_tmp.sheetnames:
            if sheet_name.upper() in TARGET_SHEETS:
                ws_tmp = wb_src_tmp[sheet_name]
                # count data rows (excluding header)
                total_rows += max(0, ws_tmp.max_row - 1)
    except Exception:
        pass

    try:
        wb_staff_tmp = load_workbook(staff_file, data_only=True)
        for sheet in TARGET_SHEETS:
            if sheet in wb_staff_tmp.sheetnames:
                ws_tmp = wb_staff_tmp[sheet]
                total_rows += max(0, ws_tmp.max_row - 1)
    except Exception:
        pass

    try:
        wb_roster_tmp = load_workbook(roster_file, data_only=True)
        for sheet_name in wb_roster_tmp.sheetnames:
            ws_tmp = wb_roster_tmp[sheet_name]
            total_rows += max(0, ws_tmp.max_row - 1)
    except Exception:
        pass

    # ensure at least 1 to avoid division by zero
    if total_rows <= 0:
        total_rows = 1

    progress_state = {'current': 0, 'total': total_rows}

    # preload staff & roster (these functions will call status_callback per row if provided)
    instructors_map = preload_staff(staff_file, status_callback=status_callback, progress_state=progress_state)
    availability, off_days = preload_roster(roster_file, status_callback=status_callback, progress_state=progress_state)

    wb_src = load_workbook(events_file, data_only=True)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    activity_cache = {}
    event_color_cache = {}
    activity_resorts = {}

    # build activity -> resorts map (NORMALIZE keys to lower-case)
    for sheet_name in wb_src.sheetnames:
        if sheet_name.upper() not in TARGET_SHEETS:
            continue
        ws = wb_src[sheet_name]
        header_map = {safe_str(ws.cell(row=1, column=c).value).lower(): c for c in range(1, ws.max_column+1)}
        activity_col = header_map.get("activity")
        resort_col = header_map.get("resort name")
        if not activity_col or not resort_col:
            continue
        # use true last data row instead of ws.max_row
        last_row = get_last_data_row(ws, activity_col)
        for r in range(2, last_row + 1):
            # per-row status update
            progress_state['current'] += 1
            if status_callback:
                status_callback(progress_state['current'], progress_state['total'], f"Scanning activities {sheet_name} row {r}")
            act = safe_str(ws.cell(row=r, column=activity_col).value)
            if not act:
                # don't break here — continue scanning further rows (months may be separated by blanks)
                continue
            res = safe_str(ws.cell(row=r, column=resort_col).value)
            act_key = act.strip().lower()
            activity_resorts.setdefault(act_key, set()).add(res)

    # Try to open xlwings book for dropdowns; if fails, continue without dropdowns
    app = xw.App(visible=False)
    try:
        wb_xlw = xw.Book(events_file)
    except Exception:
        wb_xlw = None

    for sheet_name in wb_src.sheetnames:
        if sheet_name.upper() not in TARGET_SHEETS:
            continue
        ws_src = wb_src[sheet_name]
        try:
            ws_xlw = wb_xlw.sheets[sheet_name] if wb_xlw else None
        except Exception:
            ws_xlw = None
        ws_out = wb_out.create_sheet(sheet_name)
        add_headers(ws_out)
        header_map = {safe_str(ws_src.cell(row=1, column=c).value).lower(): c for c in range(1, ws_src.max_column + 1)}
        resort_col = header_map.get("resort name")
        activity_col = header_map.get("activity")
        duration_col = header_map.get("activity duration")
        bookable_col = header_map.get("bookable hours")
        month_col = header_map.get("month")
        config_col = header_map.get("configuration")  # read configuration column if present
        product_col = header_map.get("product")  # <-- NEW: product column for GALAXEA sheet
        if month_col is None or activity_col is None or bookable_col is None:
            continue
        date_start_col = month_col + 1
        out_row = 2
        written_main_rows = set()        # track internal unique main rows
        written_instructor_rows = set()  # track instructor rows uniqueness
        activity_cache = {}

        # ---------- ROW PROCESSING ----------
        # use true last data row instead of ws_src.max_row
        last_row = get_last_data_row(ws_src, activity_col)
        for r in range(2, last_row + 1):
            # per-row status update
            progress_state['current'] += 1
            if status_callback:
                status_callback(progress_state['current'], progress_state['total'], f"Processing {sheet_name} row {r}")

            activity_raw = ws_src.cell(row=r, column=activity_col).value
            if activity_raw is None or str(activity_raw).strip() == "":
                # skip blank rows (don't break — there may be months/rows further down)
                continue

            activity = safe_str(activity_raw)
            activity_key = activity.strip().lower()

            resort = safe_str(ws_src.cell(row=r, column=resort_col).value) if resort_col else ""
            month_val = ws_src.cell(row=r, column=month_col).value

            # read configuration (preferred) — fallback to activity
            if config_col:
                config_raw = ws_src.cell(row=r, column=config_col).value
                config_for_main = safe_str(config_raw) if config_raw is not None and str(config_raw).strip() != "" else activity
            else:
                config_for_main = activity

            # special-case: ensure WAMA sailing uses "Sailing" configuration
            if sheet_name.upper() == "WAMA" and activity.lower().startswith("sailing"):
                resource_for_main = "Sailing"
                config_for_main = "Sailing"
            else:
                resource_for_main = activity

            month_num = parse_month_to_num(month_val)
            if not month_num:
                continue
            last_day = days_in_month(YEAR_FOR_OUTPUT, month_num)
            max_check_col = min(ws_src.max_column, date_start_col + last_day - 1)
            available_days = []
            for col in range(date_start_col, max_check_col + 1):
                hdr = ws_src.cell(row=1, column=col).value
                if hdr is None:
                    continue
                try:
                    day_num = int(hdr)
                except:
                    try:
                        day_num = int(float(hdr))
                    except:
                        continue
                if not (1 <= day_num <= last_day):
                    continue
                c = ws_src.cell(row=r, column=col)
                try:
                    if c.value is not None and c.value > 0:
                        capacity_val = c.value
                        available_days.append((day_num, capacity_val))
                except:
                    continue
            if not available_days:
                continue

            # fetch resorts using normalized activity key
            all_resorts_for_activity = activity_resorts.get(activity_key, set())
            event_names = build_event_names(activity, resort, all_resorts_for_activity)
            if not event_names:
                continue

            # ---------- MODIFIED: determine instr_lookup_key ----------
            # For AKUN and WAMA: keep existing logic (configuration/activity)
            # For GALAXEA: prefer Product column value (normalized) to lookup instructors
            instr_lookup_key = None
            if sheet_name.upper() == "GALAXEA" and product_col:
                product_raw = ws_src.cell(row=r, column=product_col).value
                product_val = safe_str(product_raw)
                if product_val:
                    instr_lookup_key = product_val.strip().lower()
            # fallback to configuration or activity if not set by Product (or for other sheets)
            if not instr_lookup_key:
                instr_lookup_key = config_for_main.strip().lower() if config_for_main else activity_key

            # make cache key include sheet and instr_lookup_key
            cache_key = (sheet_name.upper(), instr_lookup_key)
            if cache_key in activity_cache:
                qualified_instrs = activity_cache[cache_key]
            else:
                qualified_instrs = instructors_map.get(sheet_name.upper(), {}).get(instr_lookup_key, [])
                activity_cache[cache_key] = qualified_instrs

            resort_short = get_resort_short_from_name(resort)
            slots = []
            if ws_xlw is not None:
                try:
                    slots = get_bookable_hours(ws_xlw, r, bookable_col)
                except Exception:
                    pass
            if not slots:
                continue

            for (day_num, capacity_val) in available_days:
                date_str = f"{YEAR_FOR_OUTPUT}-{month_num:02d}-{day_num:02d}"
                roster_key = (month_num, int(day_num), resort_short) if resort_short else None
                roster_available = availability.get(roster_key, set()) if roster_key else set()
                off_for_date = off_days.get((month_num, int(day_num)), {})

                # OFF rows only for AKUN sheet
                if sheet_name.upper() == "AKUN":
                    for instr_name, off_code in off_for_date.items():
                        off_row_key = ("OFF", instr_name, date_str)
                        if off_row_key in written_instructor_rows:
                            continue
                        ws_out.cell(row=out_row, column=1, value=off_code)
                        ws_out.cell(row=out_row, column=2, value=instr_name)
                        ws_out.cell(row=out_row, column=3, value="")
                        ws_out.cell(row=out_row, column=4, value=date_str)
                        ws_out.cell(row=out_row, column=5, value="00:00:00")
                        ws_out.cell(row=out_row, column=6, value="23:59:59")
                        written_instructor_rows.add(off_row_key)
                        out_row += 1

                off_instr_names = set(off_for_date.keys())

                for event in event_names:
                    # ensure unique internal main key per resort + date (so events with same display name at different resorts don't collide)
                    internal_main_key = (event.strip().lower(), resort.strip().lower(), date_str)

                    if event not in event_color_cache:
                        event_color_cache[event] = get_light_fill()
                    fill = event_color_cache[event]

                    def normalize_time_excel(val):
                        if not val:
                            return ""
                        val = str(val).strip()
                        try:
                            if isinstance(val, datetime.time):
                                return val.strftime("%H:%M:%S")
                            try:
                                parsed = datetime.datetime.strptime(val, "%H:%M")
                            except ValueError:
                                parsed = datetime.datetime.strptime(val, "%H:%M:%S")
                            return parsed.strftime("%H:%M:%S")
                        except Exception:
                            return val

                    for slot in slots:
                        if "-" not in slot:
                            continue
                        parts = [s.strip() for s in slot.split("-", 1)]
                        if len(parts) != 2:
                            continue
                        start, end = normalize_time_excel(parts[0]), normalize_time_excel(parts[1])

                        main_key = internal_main_key
                        if main_key not in written_main_rows:
                            # resource_for_main already set above; config_for_main is the config value
                            for col_idx, val in enumerate(
                                [event, resource_for_main, config_for_main, date_str, start, end, capacity_val],
                                start=1
                            ):
                                cell = ws_out.cell(row=out_row, column=col_idx, value=val)
                                cell.fill = fill
                                cell.font = ROW_BOLD_FONT
                            month_short = MONTH_SHORT.get(month_num, f"{month_num:02d}")
                            reference_val = f"{sheet_name}-{resort}-{month_short}"
                            ws_out.cell(row=out_row, column=8, value=reference_val).fill = fill
                            ws_out.cell(row=out_row, column=8).font = ROW_BOLD_FONT
                            out_row += 1
                            written_main_rows.add(main_key)

                        # Only instructors from correct staff tab and matching configuration/product
                        instrs_to_use = [i for i in qualified_instrs if i in roster_available and i not in off_instr_names]

                        for instr in instrs_to_use:
                            # instructor uniqueness must also include resort so same instructor at different resorts/dates don't clash
                            instr_key = (event.strip().lower(), resort.strip().lower(), instr, date_str, start, end)
                            if instr_key in written_instructor_rows:
                                continue

                            # config_for_instr must be the configuration value (what we used to lookup instructors)
                            config_for_instr = config_for_main

                            for col_idx, val in enumerate([event, instr, config_for_instr, date_str, start, end], start=1):
                                cell = ws_out.cell(row=out_row, column=col_idx, value=val)
                                cell.fill = fill
                            ws_out.cell(row=out_row, column=7, value=capacity_val)
                            ws_out.cell(row=out_row, column=8, value=reference_val)
                            written_instructor_rows.add(instr_key)
                            out_row += 1

    try:
        if wb_xlw:
            wb_xlw.close()
    except:
        pass
    try:
        app.quit()
    except:
        pass
    wb_out.save(output_file)
    # final callback to reach 100%
    if status_callback:
        status_callback(progress_state['total'], progress_state['total'], "Finished generating output")
    print(f"✅ Output saved to {output_file}")


# ---------- PREVIEW ----------
def get_preview(output_file):
    excel_file = pd.ExcelFile(output_file)
    preview_dict = {}
    for sheet_name in excel_file.sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        df.insert(0, "No.", range(1, len(df) + 1))
        preview_dict[sheet_name] = df
    return preview_dict

# ---------- DATABASE PUSH ----------
def push_to_database(preview_dict, cancel_check=None):
    """
    preview_dict: {sheet_name: pandas.DataFrame}
    cancel_check: optional callable that when returns True -> cancel the push
    returns: (executed_count, errors) where errors is list of (sheet_name, row_number, error_message)
    """
    conn = None
    cursor = None
    executed_count = 0
    errors = []

    # ---------- TIME FIX: ensure start_time, end_time are datetime.time ----------
    def normalize_time(val):
        if val is None or pd.isna(val):
            return None
        if isinstance(val, datetime.time):
            return val
        if isinstance(val, datetime.datetime):
            return val.time()
        s = str(val).strip()
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.datetime.strptime(s, fmt).time()
            except Exception:
                continue
        return None

    # ---------- NEW: helpers to coerce types for SP call (no logic changes) ----------
    def _is_na(v):
        try:
            return pd.isna(v)
        except Exception:
            return False

    def to_str_or_none(v):
        if v is None or _is_na(v):
            return None
        s = str(v).strip()
        return s if s != "" else None

    def to_int_or_zero(v):
        if v is None or _is_na(v) or str(v).strip() == "":
            return 0
        try:
            # handles numpy types / floats like 5.0
            return int(float(v))
        except Exception:
            return 0

    try:
        conn = pyodbc.connect(
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=51.77.118.155\\MSSQLSERVER01;"
            "DATABASE=RSG_Unicom;"
            "UID=sysadminqa;"
            "PWD=kBmVffKYw8AM7^7y"
        )
        cursor = conn.cursor()
        cursor.fast_executemany = True

        for sheet_name, df in preview_dict.items():
            for idx, row in df.iterrows():
                if cancel_check and cancel_check():
                    if conn:
                        try:
                            conn.rollback()
                        except:
                            pass
                    errors.append((sheet_name, int(idx) + 2, "Cancelled by user"))
                    return executed_count, errors

                # ---------- EXTRACT VALUES ----------
                try:
                    event_name = row.get('Event')
                    asset_name = row.get('Resource')
                    config_name = row.get('Configuration') or ''
                    capacity = row.get('Capacity') or 0
                    date_val = row.get('Date')
                    start_time = row.get('Start Time')
                    end_time = row.get('End Time')
                    ref = row.get('Reference') or ''
                except Exception:
                    try:
                        event_name = row.iloc[0]
                        asset_name = row.iloc[1]
                        config_name = row.iloc[2] or ''
                        capacity = row.iloc[6] or 0
                        date_val = row.iloc[3]
                        start_time = row.iloc[4]
                        end_time = row.iloc[5]
                        ref = row.iloc[7] if len(row) > 7 else ''
                    except Exception as e:
                        errors.append((sheet_name, int(idx) + 2, f"Malformed row: {e}"))
                        continue

                # ---------- DATE FIX ----------
                if isinstance(date_val, str):
                    date_str_candidate = date_val.strip()
                    parsed = None
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                        try:
                            parsed = datetime.datetime.strptime(date_str_candidate, fmt).date()
                            break
                        except Exception:
                            parsed = None
                    if parsed is None:
                        try:
                            parsed = pd.to_datetime(date_str_candidate, dayfirst=True, errors="coerce")
                            if pd.isna(parsed):
                                parsed = None
                            else:
                                parsed = parsed.date()
                        except Exception:
                            parsed = None
                    date_val = parsed
                elif isinstance(date_val, pd.Timestamp):
                    date_val = date_val.date()
                elif isinstance(date_val, datetime.datetime):
                    date_val = date_val.date()
                elif isinstance(date_val, datetime.date):
                    pass
                else:
                    date_val = None

                # ---------- TIME FIX ----------
                start_time = normalize_time(start_time)
                end_time = normalize_time(end_time)

                # ---------- NEW: final coercions to SQL-friendly Python types (SP-safe) ----------
                event_name_sql = to_str_or_none(event_name)
                asset_name_sql = to_str_or_none(asset_name)
                config_name_sql = to_str_or_none(config_name)
                ref_sql = to_str_or_none(ref)

                capacity_sql = to_int_or_zero(capacity)
                inserted_user_sql = "optimo.admin"  # unchanged
                # date_val should be datetime.date or None (already normalized)
                # start_time/end_time should be datetime.time or None (already normalized)
                preliminary_sql = 0
                exclusive_sql = 0

                # ---------- DATABASE INSERT ----------
                try:
                    cursor.execute("""
                        EXEC OX_InsertEventTimeslotWithBlockingCheck
                            @EventName=?,
                            @AssetName=?,
                            @ConfigurationName=?,
                            @Capacity=?,
                            @InsertedUser=?,
                            @Date=?,
                            @StartTime=?,
                            @EndTime=?,
                            @Preliminary=?,
                            @IsSingleEventExclusive=?,
                            @Ref=?
                    """, (
                        event_name_sql,
                        asset_name_sql,
                        config_name_sql,
                        capacity_sql,
                        inserted_user_sql,
                        date_val,
                        start_time,
                        end_time,
                        int(preliminary_sql),
                        int(exclusive_sql),
                        ref_sql
                    ))
                    executed_count += 1
                except Exception as e:
                    errors.append((sheet_name, int(idx) + 2, str(e)))

        try:
            conn.commit()
        except Exception as e:
            try:
                conn.rollback()
            except:
                pass
            raise

    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except:
                pass
        raise
    finally:
        if cursor:
            try:
                cursor.close()
            except:
                pass
        if conn:
            try:
                conn.close()
            except:
                pass

    return executed_count, errors
